import os
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

‫#‬غير المسار واسم ملف إكسل حسب الموجود لديك
os.chdir('/Users/Mamj/code12')
wb = openpyxl.load_workbook('GradeBook.xlsx', read_only=True,data_only=True)

driver = webdriver.Chrome()
driver.implicitly_wait(30)
driver.get("https://noor.moe.gov.sa/")
assert "EduWave" in driver.title

user_name = "رقم بطاقتك"
password = "الرقم السري لنور"
element = driver.find_element_by_id("tbPublic")
element.send_keys(user_name)
element = driver.find_element_by_id("tbPrivate")
element.send_keys(password)

# get the submit button
bt_submit = driver.find_element_by_css_selector("[type=submit]")

# wait for the user to click the submit button (check every 1s with a 1000s timeout)
WebDriverWait(driver, timeout=300, poll_frequency=1) \
  .until(EC.staleness_of(bt_submit))

def fillMarks(ClassID, sheetName, firstStudentIDRowNumber, lastStudentIDRowNumber, shortAmlyExam, shortNdryExam, mshroaat, mosharkh, wajebat, mlfAamal, finalNdryExam, finalAmlyExam, hdoorMark, hdoorStatus):
    #الانتقال إلى صفحة اختيار الشعبة
    elem = driver.find_element_by_link_text("الدرجات").click()

    #الانتقال إلى صفحة إدخال الدرجات شعبة ٤
    elem = driver.find_element_by_id(ClassID).click()

    time.sleep(1)
    sht = wb[sheetName]

    elem = driver.find_element_by_id("ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl01_cbHeader").click()
    time.sleep(1)

    for i in range(firstStudentIDRowNumber, lastStudentIDRowNumber+1):   
        cell_obj = sht.cell(row = i, column = shortAmlyExam)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' + str(f"{i-1:02d}")+ '_tb_5')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = shortNdryExam)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' + str(f"{i-1:02d}")+ '_tb_6')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = mshroaat)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' +str(f"{i-1:02d}")+ '_tb_19')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = mosharkh)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' +str(f"{i-1:02d}")+ '_tb_22')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = wajebat)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' +str(f"{i-1:02d}")+ '_tb_24')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = mlfAamal)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' +str(f"{i-1:02d}")+ '_tb_29')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = finalNdryExam)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' +str(f"{i-1:02d}")+ '_tb_3')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = finalAmlyExam)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' +str(f"{i-1:02d}")+ '_tb_4')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = hdoorMark)
        elem = driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' +str(f"{i-1:02d}")+ '_tbStudentAttendanceMark')
        elem.clear()
        elem.send_keys(str(cell_obj.value))

        cell_obj = sht.cell(row = i, column = hdoorStatus)
        if str(cell_obj.value)== 'غائب':
            ddelement= Select(driver.find_element_by_id('ctl00_PlaceHolderMain_gvCourseSectionExamsGrades_ctl' +str(f"{i-1:02d}")+ '_ddlAttendanceStatus'))
            ddelement.select_by_visible_text('غائب')

    elem = driver.find_element_by_id("ctl00_PlaceHolderMain_ibtnSave").click()
    time.sleep(3)
# ------------------------------------------------------------------------------

‫#‬ نداء لوظيفة الرصد من خلال تحديد اسم الورقة وصف البداية والنهاية وموقع كل عمود درجات بالأرقام وليس الأحرف
fillMarks(ClassID="ctl00_PlaceHolderMain_gvCourseSections_ctl04_lbtnViewGrades", sheetName="شعبة 4", firstStudentIDRowNumber=3, lastStudentIDRowNumber=32, shortAmlyExam=5, shortNdryExam=6, mshroaat=7, mosharkh=8, wajebat=9, mlfAamal=10, finalNdryExam=11, finalAmlyExam=12, hdoorMark=13, hdoorStatus=14)
fillMarks(ClassID="ctl00_PlaceHolderMain_gvCourseSections_ctl03_lbtnViewGrades", sheetName="شعبة 5", firstStudentIDRowNumber=3, lastStudentIDRowNumber=33, shortAmlyExam=5, shortNdryExam=6, mshroaat=7, mosharkh=8, wajebat=9, mlfAamal=10, finalNdryExam=11, finalAmlyExam=12, hdoorMark=13, hdoorStatus=14)
fillMarks(ClassID="ctl00_PlaceHolderMain_gvCourseSections_ctl02_lbtnViewGrades", sheetName="شعبة 6", firstStudentIDRowNumber=3, lastStudentIDRowNumber=32, shortAmlyExam=5, shortNdryExam=6, mshroaat=7, mosharkh=8, wajebat=9, mlfAamal=10, finalNdryExam=11, finalAmlyExam=12, hdoorMark=13, hdoorStatus=14)

assert "الرابط غير صحيح أو لا يعمل." not in driver.page_source
time.sleep(99)

driver.close()
